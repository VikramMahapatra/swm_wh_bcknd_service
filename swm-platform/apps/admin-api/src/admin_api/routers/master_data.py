from __future__ import annotations

import json
import ipaddress
import math
from io import BytesIO
from datetime import date, datetime, timedelta, timezone
from typing import Any, Literal
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from pydantic import AliasChoices, BaseModel, ConfigDict, Field, field_validator, model_validator
from sqlalchemy.exc import NoResultFound
from sqlalchemy import and_, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from starlette.responses import StreamingResponse
from swm_db import (
    AssignmentCreateInput,
    DeviceORM,
    DeviceRepository,
    DeviceVehicleAssignmentORM,
    DeviceVehicleAssignmentRepository,
    DeviceVehicleAssignmentService,
    DriverORM,
    DumpYardORM,
    DumpYardWeighmentORM,
    GtsPointORM,
    GeofenceORM,
    GeofenceRepository,
    PickupPointORM,
    RouteORM,
    RouteRepository,
    SecondaryVehicleAssignmentORM,
    VehicleORM,
    VehicleRepository,
    VendorORM,
    VendorRepository,
    WardORM,
    WardRepository,
    ZoneORM,
    get_db_session,
)

from admin_api.api_support import (
    MessageResponse,
    PageResponse,
    RoleContext,
    _fetch_or_404,
    _list_entities,
    _parse_bool,
    _parse_csv_with_required,
    _parse_uuid,
    _raise_not_found,
    _to_dict,
    require_roles,
)

router = APIRouter()

SECONDARY_WASTE_TYPES = [
    {"value": "chicken_waste", "label": "CHICKEN WASTE"},
    {"value": "biomedical_waste", "label": "BIOMEDICAL WASTE"},
    {"value": "construction_waste", "label": "CONSTRUCTION WASTE"},
    {"value": "dry_waste", "label": "DRY WASTE"},
    {"value": "green_waste", "label": "GREEN WASTE"},
    {"value": "mandai", "label": "MANDAI"},
    {"value": "mix_waste", "label": "MIX WASTE"},
    {"value": "mixed_waste", "label": "MIXED WASTE"},
    {"value": "plastic_waste", "label": "PLASTIC WASTE"},
    {"value": "wet_waste", "label": "WET WASTE"},
]

SECONDARY_WASTE_TYPE_VALUES = {item["value"] for item in SECONDARY_WASTE_TYPES}


def _parse_import_uuid(value: str | None, *, field: str, row_number: int, required: bool = True) -> UUID | None:
    normalized = (value or "").strip()
    if not normalized:
        if required:
            raise HTTPException(status_code=400, detail=f"invalid {field} at row {row_number}")
        return None
    try:
        return UUID(normalized)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"invalid {field} at row {row_number}") from exc


def _parse_import_float(value: str | None, *, field: str, row_number: int, default: float | None = None) -> float | None:
    normalized = (value or "").strip()
    if not normalized:
        return default
    try:
        return float(normalized)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"invalid {field} at row {row_number}") from exc


def _parse_import_int(value: str | None, *, field: str, row_number: int, default: int | None = None) -> int | None:
    normalized = (value or "").strip()
    if not normalized:
        return default
    try:
        return int(normalized)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"invalid {field} at row {row_number}") from exc


def _parse_import_datetime(value: str | None, *, field: str, row_number: int) -> datetime | None:
    normalized = (value or "").strip()
    if not normalized:
        return None
    try:
        return datetime.fromisoformat(normalized)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"invalid {field} at row {row_number}") from exc


class VendorIn(BaseModel):
    vendor_code: str = Field(min_length=3, max_length=32, pattern=r"^[A-Z0-9_-]{3,32}$")
    vendor_name: str = Field(min_length=1)
    contact_person: str | None = None
    email: str | None = None
    phone: str | None = None
    gst_number: str | None = None
    address: str | None = None
    contract_start: str | None = None
    contract_end: str | None = None
    webhook_secret: str | None = None
    signature_key: str | None = None
    allowed_ips: list[str] = Field(default_factory=list)
    auth_type: Literal["header", "signature", "ip"] = "header"
    callback_format: dict[str, Any] = Field(default_factory=dict)
    active: bool = True
    metadata: dict[str, Any] = Field(default_factory=dict)

    @field_validator("email")
    @classmethod
    def validate_email(cls, value: str | None) -> str | None:
        if value is None:
            return None
        normalized = value.strip().lower()
        if "@" not in normalized or normalized.startswith("@") or normalized.endswith("@"):
            raise ValueError("email is not valid")
        local_part, _, domain = normalized.partition("@")
        if not local_part or "." not in domain or domain.startswith(".") or domain.endswith("."):
            raise ValueError("email is not valid")
        return normalized

    @field_validator("allowed_ips")
    @classmethod
    def validate_allowed_ips(cls, value: list[str]) -> list[str]:
        normalized: list[str] = []
        for ip in value:
            normalized.append(str(ipaddress.ip_address(ip)))
        return normalized

    @field_validator("phone")
    @classmethod
    def validate_phone_required(cls, value: str | None) -> str:
        normalized = (value or "").strip()
        if not normalized:
            raise ValueError("phone is required")
        return normalized

    @model_validator(mode="after")
    def validate_contract_period(self) -> "VendorIn":
        start = (self.contract_start or "").strip()
        end = (self.contract_end or "").strip()
        if start and end:
          try:
              start_date = date.fromisoformat(start)
              end_date = date.fromisoformat(end)
          except ValueError as exc:
              raise ValueError("contract dates must use YYYY-MM-DD format") from exc
          if end_date < start_date:
              raise ValueError("contract_end must be greater than or equal to contract_start")
        return self


def _vendor_metadata(
    metadata: dict[str, Any] | None,
    contract_start: str | None,
    contract_end: str | None,
    gst_number: str | None,
    address: str | None,
) -> dict[str, Any]:
    payload = dict(metadata or {})
    if contract_start is not None:
        payload["contract_start"] = contract_start
    if contract_end is not None:
        payload["contract_end"] = contract_end
    if gst_number is not None:
        payload["gst_number"] = gst_number
    if address is not None:
        payload["address"] = address
    return payload


def _vendor_response(vendor: dict[str, Any]) -> dict[str, Any]:
    metadata = vendor.get("metadata") if isinstance(vendor.get("metadata"), dict) else {}
    contract_start = metadata.get("contract_start")
    contract_end = metadata.get("contract_end")
    gst_number = metadata.get("gst_number")
    address = metadata.get("address")
    vendor["contract_start"] = contract_start
    vendor["contract_end"] = contract_end
    vendor["contractStart"] = contract_start
    vendor["contractEnd"] = contract_end
    vendor["gst_number"] = gst_number
    vendor["gstNumber"] = gst_number
    vendor["address"] = address
    return vendor


class DeviceIn(BaseModel):
    vendor_id: UUID
    imei: str = Field(min_length=14, max_length=17, pattern=r"^[0-9]{14,17}$")
    serial_no: str | None = None
    model: str | None = None
    manufacturer: str | None = None
    firmware_version: str | None = None
    sim_number: str | None = None
    installed_on: datetime | None = None
    activated_on: datetime | None = None
    last_seen: datetime | None = None
    battery_percent: float | None = Field(default=None, ge=0, le=100)
    signal_strength: float | None = None
    health_status: Literal["healthy", "warning", "critical", "offline"] = "healthy"
    active: bool = True
    metadata: dict[str, Any] = Field(default_factory=dict)


class VehicleIn(BaseModel):
    model_config = ConfigDict(populate_by_name=True)

    vehicle_number: str = Field(min_length=4, max_length=24, pattern=r"^[A-Z0-9-]{4,24}$")
    registration_number: str = Field(min_length=4, max_length=24, pattern=r"^[A-Z0-9-]{4,24}$")
    vendor_id: UUID = Field(validation_alias=AliasChoices("vendor_id", "vendorId", "contractor_id", "contractorId"))
    ward_id: UUID
    route_id: UUID | None = None
    vehicle_category: Literal["primary", "secondary"] = "primary"
    secondary_waste_type: Literal[
        "chicken_waste",
        "dry_waste",
        "green_waste",
        "mandai",
        "mix_waste",
        "wet_waste",
    ] | None = None
    truck_type: str | None = None
    capacity_kg: float = Field(default=0, ge=0)
    capacity_cubic_meter: float = Field(default=0, ge=0)
    fuel_type: Literal["diesel", "petrol", "cng", "electric", "lng"] = "diesel"
    operational_status: Literal["operational", "maintenance", "breakdown", "retired"] = "operational"
    chassis_number: str | None = None
    engine_number: str | None = None
    manufacture_year: int | None = Field(default=None, ge=1950, le=2100)
    active: bool = True
    metadata: dict[str, Any] = Field(default_factory=dict)


class DumpYardIn(BaseModel):
    dump_yard_code: str | None = Field(default=None, min_length=2, max_length=32)
    dump_yard_name: str | None = Field(default=None, min_length=1)
    name: str | None = None
    address: str | None = None
    capacity: float | None = Field(default=None, ge=0)
    zone_id: UUID | None = None
    ward_id: UUID | None = None
    lat: float | None = Field(default=None, ge=-90, le=90)
    lng: float | None = Field(default=None, ge=-180, le=180)
    latitude: float | None = Field(default=None, ge=-90, le=90)
    longitude: float | None = Field(default=None, ge=-180, le=180)
    active: bool = True
    is_active: bool | None = None


class GtsIn(BaseModel):
    name: str = Field(min_length=1)
    latitude: float | None = Field(default=None, ge=-90, le=90)
    longitude: float | None = Field(default=None, ge=-180, le=180)
    address: str | None = None
    zone_id: UUID | None = None
    ward_id: UUID | None = None
    is_active: bool = True


class SecondaryVehicleAssignmentIn(BaseModel):
    vehicle_id: UUID
    gtc_pickup_point_id: UUID
    dump_yard_id: UUID
    material_type: Literal[
        "chicken_waste",
        "biomedical_waste",
        "construction_waste",
        "dry_waste",
        "green_waste",
        "mandai",
        "mix_waste",
        "mixed_waste",
        "plastic_waste",
        "wet_waste",
    ]
    assigned_from: datetime | None = None
    assigned_to: datetime | None = None
    active: bool = True
    remarks: str | None = None


class DumpYardWeighmentIn(BaseModel):
    assignment_id: UUID | None = None
    vehicle_id: UUID
    gtc_pickup_point_id: UUID | None = Field(
        default=None,
        validation_alias=AliasChoices("gtc_pickup_point_id", "gts_pickup_point_id", "GTS_pickup_point_id"),
    )
    dump_yard_id: UUID
    material_type: Literal[
        "chicken_waste",
        "biomedical_waste",
        "construction_waste",
        "dry_waste",
        "green_waste",
        "mandai",
        "mix_waste",
        "mixed_waste",
        "plastic_waste",
        "wet_waste",
    ]
    service_date: date | None = None
    entry_time: datetime | None = None
    gross_weight_kg: float = Field(ge=0)
    tare_weight_kg: float = Field(ge=0)
    net_weight_kg: float | None = Field(default=None, ge=0)
    slip_number: str | None = None
    operator_name: str | None = None
    remarks: str | None = None


class WardIn(BaseModel):
    ward_code: str
    ward_name: str
    zone_name: str
    active: bool = True
    population: int | None = None
    area: float | None = None
    total_pickup_points: int | None = None


async def _resolve_zone_id(session: AsyncSession, zone_ref: str) -> UUID:
    normalized = zone_ref.strip()
    if not normalized:
        raise HTTPException(status_code=400, detail="zone_name is required")

    try:
        zone_uuid = UUID(normalized)
        zone = (await session.execute(select(ZoneORM).where(ZoneORM.id == zone_uuid))).scalars().first()
        if zone is not None:
            return zone.id
    except ValueError:
        pass

    zone = (
        await session.execute(
            select(ZoneORM).where(
                (func.lower(ZoneORM.zone_code) == normalized.lower())
                | (func.lower(ZoneORM.zone_name) == normalized.lower())
            )
        )
    ).scalars().first()
    if zone is None:
        raise HTTPException(status_code=400, detail=f"zone not found for '{zone_ref}'")
    return zone.id


class RouteIn(BaseModel):
    route_name: str = Field(min_length=1)
    route_type: Literal["primary", "secondary"] = "primary"
    zone_id: UUID
    ward_id: UUID
    polyline_coordinates: list[list[float]] = Field(min_length=2)


def _distance_km_between(
    first: tuple[float | None, float | None],
    second: tuple[float | None, float | None],
) -> float:
    if first[0] is None or first[1] is None or second[0] is None or second[1] is None:
        return 0.0
    lat1, lng1 = float(first[0]), float(first[1])
    lat2, lng2 = float(second[0]), float(second[1])
    radius_km = 6371.0
    d_lat = math.radians(lat2 - lat1)
    d_lng = math.radians(lng2 - lng1)
    a = (
        math.sin(d_lat / 2) ** 2
        + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(d_lng / 2) ** 2
    )
    return radius_km * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def _minutes_from_hhmm(value: str | None) -> int | None:
    if not value:
        return None
    try:
        hours, minutes = str(value).strip()[:5].split(":")
        return int(hours) * 60 + int(minutes)
    except Exception:
        return None


def _route_stats_from_points(points: list[PickupPointORM]) -> dict[str, Any]:
    ordered = sorted(points, key=lambda p: (p.sequence_no or 0, str(p.id)))
    distance_km = 0.0
    for first, second in zip(ordered, ordered[1:]):
        distance_km += _distance_km_between((first.lat, first.lng), (second.lat, second.lng))

    times = [_minutes_from_hhmm(point.expected_pickup_time) for point in ordered]
    valid_times = [value for value in times if value is not None]
    if len(valid_times) >= 2:
        estimated_minutes = max(0, max(valid_times) - min(valid_times))
    else:
        estimated_minutes = int(round((distance_km / 14.0) * 60 + len(ordered) * 5)) if ordered else 0

    return {
        "total_pickup_points": len(ordered),
        "totalPickupPoints": len(ordered),
        "estimated_distance": round(distance_km, 2),
        "estimatedDistance": round(distance_km, 2),
        "distance": f"{distance_km:.1f} km",
        "estimated_time": estimated_minutes,
        "estimatedTime": estimated_minutes,
        "has_gts": any(bool(getattr(point, "is_gts", False)) for point in ordered),
        "hasGts": any(bool(getattr(point, "is_gts", False)) for point in ordered),
    }


class GeofenceIn(BaseModel):
    geofence_code: str = Field(min_length=2, max_length=32, pattern=r"^[A-Z0-9_-]{2,32}$")
    geofence_name: str = Field(min_length=1)
    type: Literal["depot", "landfill", "zone", "parking", "maintenance"]
    geometry_type: Literal["circle", "polygon"]
    center_lat: float | None = Field(default=None, ge=-90, le=90)
    center_lng: float | None = Field(default=None, ge=-180, le=180)
    radius_meter: float | None = Field(default=None, gt=0)
    polygon: dict[str, Any] | None = None
    geofence_for: Literal["zone", "ward", "route"] = "ward"
    zone_id: UUID
    ward_id: UUID | None = None
    route_id: UUID | None = None
    active: bool = True

    @field_validator("polygon")
    @classmethod
    def validate_polygon(cls, value: dict[str, Any] | None) -> dict[str, Any] | None:
        if value is None:
            return None
        if value.get("type") != "Polygon":
            raise ValueError("polygon GeoJSON type must be Polygon")
        coordinates = value.get("coordinates")
        if not isinstance(coordinates, list) or len(coordinates) == 0:
            raise ValueError("polygon.coordinates must be a non-empty array")
        return value


class DeviceAssignmentIn(BaseModel):
    model_config = ConfigDict(populate_by_name=True)

    device_id: UUID
    vehicle_id: UUID
    assigned_from: datetime | None = None
    assigned_to: datetime | None = None
    active: bool = True
    remarks: str | None = None


# --- Zone CRUD Endpoints ---
class ZoneIn(BaseModel):
    zone_code: str = Field(min_length=2, max_length=32, pattern=r"^[A-Z0-9_-]{2,32}$")
    zone_name: str = Field(min_length=1)
    description: str | None = None
    supervisor_name: str | None = None
    supervisor_phone: str | None = None
    active: bool = True

class ZoneOut(BaseModel):
    id: UUID
    zone_code: str
    zone_name: str
    description: str | None = None
    supervisor_name: str | None = None
    supervisor_phone: str | None = None
    active: bool

@router.post("/zones", response_model=ZoneOut)
async def create_zone(payload: ZoneIn, session: AsyncSession = Depends(get_db_session)):
    zone = ZoneORM(
        zone_code=payload.zone_code,
        zone_name=payload.zone_name,
        description=payload.description,
        supervisor_name=payload.supervisor_name,
        supervisor_phone=payload.supervisor_phone,
        active=payload.active,
    )
    session.add(zone)
    await session.commit()
    await session.refresh(zone)
    return zone

@router.get("/zones", response_model=list[ZoneOut])
async def list_zones(session: AsyncSession = Depends(get_db_session)):
    result = await session.execute(select(ZoneORM).order_by(ZoneORM.zone_code))
    return result.scalars().all()

@router.get("/zones/{zone_id}", response_model=ZoneOut)
async def get_zone(zone_id: UUID, session: AsyncSession = Depends(get_db_session)):
    zone = await session.get(ZoneORM, zone_id)
    if not zone:
        raise HTTPException(status_code=404, detail="Zone not found")
    return zone

@router.put("/zones/{zone_id}", response_model=ZoneOut)
async def update_zone(zone_id: UUID, payload: ZoneIn, session: AsyncSession = Depends(get_db_session)):
    zone = await session.get(ZoneORM, zone_id)
    if not zone:
        raise HTTPException(status_code=404, detail="Zone not found")
    zone.zone_code = payload.zone_code
    zone.zone_name = payload.zone_name
    zone.description = payload.description
    zone.supervisor_name = payload.supervisor_name
    zone.supervisor_phone = payload.supervisor_phone
    zone.active = payload.active
    await session.commit()
    await session.refresh(zone)
    return zone

@router.delete("/zones/{zone_id}", response_model=MessageResponse)
async def delete_zone(zone_id: UUID, session: AsyncSession = Depends(get_db_session)):
    zone = await session.get(ZoneORM, zone_id)
    if not zone:
        raise HTTPException(status_code=404, detail="Zone not found")
    await session.delete(zone)
    await session.commit()
    return MessageResponse(message="Zone deleted successfully")


@router.get("/vendors")
async def list_vendors(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=200),
    q: str | None = Query(default=None),
    active: bool | None = Query(default=None),
    auth_type: str | None = Query(default=None),
    ui_compat: bool = Query(default=False),
    sort_by: str = Query(default="created_at"),
    sort_order: str = Query(default="desc", pattern="^(asc|desc)$"),
    _: RoleContext = Depends(require_roles("admin", "ops", "viewer")),
    session: AsyncSession = Depends(get_db_session),
) -> PageResponse | list[dict[str, Any]]:
    result = await _list_entities(
        session,
        VendorORM,
        page=page,
        page_size=page_size,
        q=q,
        sort_by=sort_by,
        sort_order=sort_order,
        filters={"active": active, "auth_type": auth_type},
    )

    if not ui_compat:
        return result

    items: list[dict[str, Any]] = []
    for vendor in result.items:
        item_id = vendor.get("id")
        metadata = vendor.get("metadata") if isinstance(vendor.get("metadata"), dict) else {}
        items.append(
            _vendor_response(
                {
                    "id": str(item_id) if item_id is not None else "",
                    "name": vendor.get("contact_person") or "",
                    "contact_person": vendor.get("contact_person") or "",
                    "company_name": vendor.get("vendor_name") or "",
                    "companyName": vendor.get("vendor_name") or "",
                    "phone": vendor.get("phone"),
                    "email": vendor.get("email"),
                    "status": "active" if vendor.get("active", True) else "inactive",
                    "active": vendor.get("active", True),
                    "metadata": metadata,
                }
            )
        )
    return items


@router.post("/vendors")
async def create_vendor(
    payload: VendorIn,
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    repo = VendorRepository(session)
    row = await repo.create(
        vendor_code=payload.vendor_code,
        vendor_name=payload.vendor_name,
        contact_person=payload.contact_person,
        email=payload.email,
        phone=payload.phone,
        webhook_secret=payload.webhook_secret,
        signature_key=payload.signature_key,
        allowed_ips=payload.allowed_ips,
        auth_type=payload.auth_type,
        callback_format=payload.callback_format,
        active=payload.active,
        metadata_json=_vendor_metadata(
            payload.metadata,
            payload.contract_start,
            payload.contract_end,
            payload.gst_number,
            payload.address,
        ),
    )
    return _vendor_response(_to_dict(row))


@router.get("/vendors/{vendor_id}")
async def get_vendor(
    vendor_id: UUID,
    _: RoleContext = Depends(require_roles("admin", "ops", "viewer")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    repo = VendorRepository(session)
    row = await _fetch_or_404(repo.get_by_id, "vendor", vendor_id)
    return _vendor_response(_to_dict(row))


@router.put("/vendors/{vendor_id}")
async def update_vendor(
    vendor_id: UUID,
    payload: VendorIn,
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    repo = VendorRepository(session)
    try:
        row = await repo.update(
            vendor_id,
            vendor_code=payload.vendor_code,
            vendor_name=payload.vendor_name,
            contact_person=payload.contact_person,
            email=payload.email,
            phone=payload.phone,
            webhook_secret=payload.webhook_secret,
            signature_key=payload.signature_key,
            allowed_ips=payload.allowed_ips,
            auth_type=payload.auth_type,
            callback_format=payload.callback_format,
            active=payload.active,
            metadata_json=_vendor_metadata(
                payload.metadata,
                payload.contract_start,
                payload.contract_end,
                payload.gst_number,
                payload.address,
            ),
        )
    except NoResultFound:
        _raise_not_found("vendor", vendor_id)
    return _vendor_response(_to_dict(row))


@router.delete("/vendors/{vendor_id}", response_model=MessageResponse)
async def delete_vendor(
    vendor_id: UUID,
    _: RoleContext = Depends(require_roles("admin")),
    session: AsyncSession = Depends(get_db_session),
) -> MessageResponse:
    repo = VendorRepository(session)
    try:
        await repo.delete(vendor_id)
    except NoResultFound:
        _raise_not_found("vendor", vendor_id)
    return MessageResponse(message="deleted")


@router.post("/vendors/import")
async def bulk_import_vendors(
    file: UploadFile = File(...),
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    content = (await file.read()).decode("utf-8")
    rows = _parse_csv_with_required(content, required_columns={"vendor_code", "vendor_name"})
    repo = VendorRepository(session)
    created = await repo.bulk_create(
        [
            {
                "vendor_code": r["vendor_code"],
                "vendor_name": r["vendor_name"],
                "contact_person": r.get("contact_person") or None,
                "email": r.get("email") or None,
                "phone": r.get("phone") or None,
                "webhook_secret": r.get("webhook_secret") or None,
                "signature_key": r.get("signature_key") or None,
                "allowed_ips": [ip.strip() for ip in (r.get("allowed_ips") or "").split(";") if ip.strip()],
                "auth_type": r.get("auth_type") or "header",
                "callback_format": {},
                "active": _parse_bool(r.get("active"), default=True),
                "metadata_json": {},
            }
            for r in rows
        ]
    )
    return {"created": len(created)}


@router.get("/devices", response_model=PageResponse)
async def list_devices(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=200),
    q: str | None = Query(default=None),
    active: bool | None = Query(default=None),
    vendor_id: UUID | None = Query(default=None),
    health_status: str | None = Query(default=None),
    sort_by: str = Query(default="created_at"),
    sort_order: str = Query(default="desc", pattern="^(asc|desc)$"),
    _: RoleContext = Depends(require_roles("admin", "ops", "viewer")),
    session: AsyncSession = Depends(get_db_session),
) -> PageResponse:
    return await _list_entities(
        session,
        DeviceORM,
        page=page,
        page_size=page_size,
        q=q,
        sort_by=sort_by,
        sort_order=sort_order,
        filters={"active": active, "vendor_id": vendor_id, "health_status": health_status},
    )


@router.post("/devices")
async def create_device(
    payload: DeviceIn,
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    repo = DeviceRepository(session)
    row = await repo.create(
        vendor_id=payload.vendor_id,
        imei=payload.imei,
        serial_no=payload.serial_no,
        model=payload.model,
        manufacturer=payload.manufacturer,
        firmware_version=payload.firmware_version,
        sim_number=payload.sim_number,
        installed_on=payload.installed_on,
        activated_on=payload.activated_on,
        last_seen=payload.last_seen,
        battery_percent=payload.battery_percent,
        signal_strength=payload.signal_strength,
        health_status=payload.health_status,
        active=payload.active,
        metadata_json=payload.metadata,
    )
    return _to_dict(row)


@router.get("/devices/{device_id}")
async def get_device(
    device_id: UUID,
    _: RoleContext = Depends(require_roles("admin", "ops", "viewer")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    repo = DeviceRepository(session)
    row = await _fetch_or_404(repo.get_by_id, "device", device_id)
    return _to_dict(row)


@router.put("/devices/{device_id}")
async def update_device(
    device_id: UUID,
    payload: DeviceIn,
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    repo = DeviceRepository(session)
    try:
        row = await repo.update(
            device_id,
            vendor_id=payload.vendor_id,
            imei=payload.imei,
            serial_no=payload.serial_no,
            model=payload.model,
            manufacturer=payload.manufacturer,
            firmware_version=payload.firmware_version,
            sim_number=payload.sim_number,
            installed_on=payload.installed_on,
            activated_on=payload.activated_on,
            last_seen=payload.last_seen,
            battery_percent=payload.battery_percent,
            signal_strength=payload.signal_strength,
            health_status=payload.health_status,
            active=payload.active,
            metadata_json=payload.metadata,
        )
    except NoResultFound:
        _raise_not_found("device", device_id)
    return _to_dict(row)


@router.delete("/devices/{device_id}", response_model=MessageResponse)
async def delete_device(
    device_id: UUID,
    _: RoleContext = Depends(require_roles("admin")),
    session: AsyncSession = Depends(get_db_session),
) -> MessageResponse:
    repo = DeviceRepository(session)
    try:
        await repo.delete(device_id)
    except NoResultFound:
        _raise_not_found("device", device_id)
    return MessageResponse(message="deleted")


@router.post("/devices/import")
async def bulk_import_devices(
    file: UploadFile = File(...),
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    content = (await file.read()).decode("utf-8")
    rows = _parse_csv_with_required(content, required_columns={"vendor_id", "imei"})
    repo = DeviceRepository(session)
    payloads: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=2):
        payloads.append(
            {
                "vendor_id": _parse_import_uuid(row.get("vendor_id"), field="vendor_id", row_number=index),
                "imei": row["imei"],
                "serial_no": row.get("serial_no") or None,
                "model": row.get("model") or None,
                "manufacturer": row.get("manufacturer") or None,
                "firmware_version": row.get("firmware_version") or None,
                "sim_number": row.get("sim_number") or None,
                "health_status": row.get("health_status") or "healthy",
                "active": _parse_bool(row.get("active"), default=True),
                "metadata_json": {},
            }
        )
    created = await repo.bulk_create(payloads)
    return {"created": len(created)}


@router.get("/vehicles", response_model=PageResponse)
async def list_vehicles(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=200),
    q: str | None = Query(default=None),
    active: bool | None = Query(default=None),
    vendor_id: UUID | None = Query(default=None),
    ward_id: UUID | None = Query(default=None),
    route_id: UUID | None = Query(default=None),
    fuel_type: str | None = Query(default=None),
    operational_status: str | None = Query(default=None),
    vehicle_category: str | None = Query(default=None),
    sort_by: str = Query(default="created_at"),
    sort_order: str = Query(default="desc", pattern="^(asc|desc)$"),
    _: RoleContext = Depends(require_roles("admin", "ops", "viewer")),
    session: AsyncSession = Depends(get_db_session),
) -> PageResponse:
    return await _list_entities(
        session,
        VehicleORM,
        page=page,
        page_size=page_size,
        q=q,
        sort_by=sort_by,
        sort_order=sort_order,
        filters={
            "active": active,
            "vendor_id": vendor_id,
            "ward_id": ward_id,
            "route_id": route_id,
            "fuel_type": fuel_type,
            "operational_status": operational_status,
            "vehicle_category": vehicle_category,
        },
    )


@router.post("/vehicles")
async def create_vehicle(
    payload: VehicleIn,
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    if payload.vehicle_category == "secondary" and not payload.secondary_waste_type:
        raise HTTPException(status_code=422, detail="secondary_waste_type is required for secondary vehicles")
    repo = VehicleRepository(session)
    row = await repo.create(
        vehicle_number=payload.vehicle_number,
        registration_number=payload.registration_number,
        vehicle_category=payload.vehicle_category,
        secondary_waste_type=payload.secondary_waste_type if payload.vehicle_category == "secondary" else None,
        truck_type=payload.truck_type,
        capacity_kg=payload.capacity_kg,
        capacity_cubic_meter=payload.capacity_cubic_meter,
        vendor_id=payload.vendor_id,
        ward_id=payload.ward_id,
        route_id=None if payload.vehicle_category == "secondary" else payload.route_id,
        fuel_type=payload.fuel_type,
        operational_status=payload.operational_status,
        chassis_number=payload.chassis_number,
        engine_number=payload.engine_number,
        manufacture_year=payload.manufacture_year,
        active=payload.active,
        metadata_json=payload.metadata,
    )
    return _to_dict(row)


def _material_label(value: str | None) -> str | None:
    if not value:
        return None
    labels = {item["value"]: item["label"] for item in SECONDARY_WASTE_TYPES}
    return labels.get(str(value), str(value).replace("_", " ").upper())


def _vehicle_search_item(
    vehicle: VehicleORM,
    *,
    ward: WardORM | None = None,
    zone: ZoneORM | None = None,
    route: RouteORM | None = None,
) -> dict[str, Any]:
    return {
        "id": str(vehicle.id),
        "vehicle_id": str(vehicle.id),
        "vehicle_number": vehicle.vehicle_number,
        "registration_number": vehicle.registration_number,
        "label": f"{vehicle.vehicle_number} | {vehicle.registration_number}",
        "vehicle_category": vehicle.vehicle_category,
        "vehicle_type": vehicle.truck_type,
        "truck_type": vehicle.truck_type,
        "material_type": vehicle.secondary_waste_type,
        "material_label": _material_label(vehicle.secondary_waste_type),
        "route_id": str(route.id) if route else (str(vehicle.route_id) if vehicle.route_id else None),
        "route_name": getattr(route, "route_name", None),
        "ward_id": str(ward.id) if ward else str(vehicle.ward_id),
        "ward_name": getattr(ward, "ward_name", None),
        "zone_id": str(zone.id) if zone else None,
        "zone_name": getattr(zone, "zone_name", None),
        "active": vehicle.active,
    }


@router.get("/vehicles/search")
async def search_vehicles(
    q: str = Query(default="", min_length=0),
    page_size: int = Query(default=20, ge=1, le=50),
    _: RoleContext = Depends(require_roles("admin", "ops", "viewer")),
    session: AsyncSession = Depends(get_db_session),
) -> list[dict[str, Any]]:
    pattern = f"%{q.strip()}%"
    stmt = (
        select(VehicleORM, WardORM, ZoneORM, RouteORM)
        .join(WardORM, VehicleORM.ward_id == WardORM.id)
        .join(ZoneORM, WardORM.zone_id == ZoneORM.id)
        .outerjoin(RouteORM, VehicleORM.route_id == RouteORM.id)
        .where(VehicleORM.active.is_(True))
        .order_by(VehicleORM.vehicle_number.asc())
        .limit(page_size)
    )
    if q.strip():
        stmt = stmt.where(
            or_(
                VehicleORM.vehicle_number.ilike(pattern),
                VehicleORM.registration_number.ilike(pattern),
                RouteORM.route_name.ilike(pattern),
                WardORM.ward_name.ilike(pattern),
                ZoneORM.zone_name.ilike(pattern),
            )
        )
    rows = (await session.execute(stmt)).all()
    return [_vehicle_search_item(vehicle, ward=ward, zone=zone, route=route) for vehicle, ward, zone, route in rows]


@router.get("/vehicles/{vehicle_id}/details")
async def get_vehicle_details(
    vehicle_id: UUID,
    _: RoleContext = Depends(require_roles("admin", "ops", "viewer")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    result = await session.execute(
        select(VehicleORM, WardORM, ZoneORM, RouteORM)
        .join(WardORM, VehicleORM.ward_id == WardORM.id)
        .join(ZoneORM, WardORM.zone_id == ZoneORM.id)
        .outerjoin(RouteORM, VehicleORM.route_id == RouteORM.id)
        .where(VehicleORM.id == vehicle_id)
    )
    row = result.first()
    if row is None:
        _raise_not_found("vehicle", vehicle_id)
    vehicle, ward, zone, route = row
    driver = (
        await session.execute(
            select(DriverORM)
            .where(
                DriverORM.assigned_vehicle_id == vehicle.id,
                DriverORM.active.is_(True),
                DriverORM.person_type == "driver",
            )
            .order_by(DriverORM.updated_at.desc())
        )
    ).scalars().first()
    assignment = (
        await session.execute(
            select(SecondaryVehicleAssignmentORM)
            .where(SecondaryVehicleAssignmentORM.vehicle_id == vehicle.id, SecondaryVehicleAssignmentORM.active.is_(True))
            .order_by(SecondaryVehicleAssignmentORM.assigned_from.desc())
        )
    ).scalars().first()
    gts_pickup = None
    dump_yard = None
    if assignment is not None:
        await session.refresh(assignment, attribute_names=["gtc_pickup_point", "dump_yard"])
        gts_pickup = assignment.gtc_pickup_point
        dump_yard = assignment.dump_yard
    if gts_pickup is None and vehicle.route_id is not None:
        gts_pickup = (
            await session.execute(
                select(PickupPointORM)
                .where(PickupPointORM.route_id == vehicle.route_id, PickupPointORM.is_gts.is_(True))
                .order_by(PickupPointORM.sequence_no.desc())
            )
        ).scalars().first()
    if dump_yard is None:
        dump_yard = (
            await session.execute(
                select(DumpYardORM)
                .where(DumpYardORM.active.is_(True))
                .order_by(DumpYardORM.dump_yard_name.asc())
            )
        ).scalars().first()

    base = _vehicle_search_item(vehicle, ward=ward, zone=zone, route=route)
    base.update(
        {
            "driver_id": str(driver.id) if driver else None,
            "driver_name": getattr(driver, "name", None),
            "last_known_route": getattr(route, "route_name", None),
            "current_assignment": {
                "id": str(assignment.id),
                "assigned_from": assignment.assigned_from.isoformat() if assignment.assigned_from else None,
                "assigned_to": assignment.assigned_to.isoformat() if assignment.assigned_to else None,
                "remarks": assignment.remarks,
            }
            if assignment
            else None,
            "gts": {
                "id": str(gts_pickup.id),
                "name": gts_pickup.pickup_name,
                "latitude": gts_pickup.lat,
                "longitude": gts_pickup.lng,
            }
            if gts_pickup
            else None,
            "gts_pickup_point_id": str(gts_pickup.id) if gts_pickup else None,
            "dump_yard": {
                "id": str(dump_yard.id),
                "name": dump_yard.dump_yard_name,
                "latitude": dump_yard.lat,
                "longitude": dump_yard.lng,
            }
            if dump_yard
            else None,
            "dump_yard_id": str(dump_yard.id) if dump_yard else None,
        }
    )
    return base


@router.get("/vehicles/{vehicle_id}")
async def get_vehicle(
    vehicle_id: UUID,
    _: RoleContext = Depends(require_roles("admin", "ops", "viewer")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    repo = VehicleRepository(session)
    row = await _fetch_or_404(repo.get_by_id, "vehicle", vehicle_id)
    return _to_dict(row)


@router.put("/vehicles/{vehicle_id}")
async def update_vehicle(
    vehicle_id: UUID,
    payload: VehicleIn,
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    if payload.vehicle_category == "secondary" and not payload.secondary_waste_type:
        raise HTTPException(status_code=422, detail="secondary_waste_type is required for secondary vehicles")
    repo = VehicleRepository(session)
    try:
        row = await repo.update(
            vehicle_id,
            vehicle_number=payload.vehicle_number,
            registration_number=payload.registration_number,
            vehicle_category=payload.vehicle_category,
            secondary_waste_type=payload.secondary_waste_type if payload.vehicle_category == "secondary" else None,
            truck_type=payload.truck_type,
            capacity_kg=payload.capacity_kg,
            capacity_cubic_meter=payload.capacity_cubic_meter,
            vendor_id=payload.vendor_id,
            ward_id=payload.ward_id,
            route_id=None if payload.vehicle_category == "secondary" else payload.route_id,
            fuel_type=payload.fuel_type,
            operational_status=payload.operational_status,
            chassis_number=payload.chassis_number,
            engine_number=payload.engine_number,
            manufacture_year=payload.manufacture_year,
            active=payload.active,
            metadata_json=payload.metadata,
        )
    except NoResultFound:
        _raise_not_found("vehicle", vehicle_id)
    return _to_dict(row)


@router.delete("/vehicles/{vehicle_id}", response_model=MessageResponse)
async def delete_vehicle(
    vehicle_id: UUID,
    _: RoleContext = Depends(require_roles("admin")),
    session: AsyncSession = Depends(get_db_session),
) -> MessageResponse:
    repo = VehicleRepository(session)
    try:
        await repo.delete(vehicle_id)
    except NoResultFound:
        _raise_not_found("vehicle", vehicle_id)
    return MessageResponse(message="deleted")


@router.post("/vehicles/import")
async def bulk_import_vehicles(
    file: UploadFile = File(...),
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    content = (await file.read()).decode("utf-8")
    rows = _parse_csv_with_required(
        content,
        required_columns={"vehicle_number", "registration_number", "vendor_id", "ward_id"},
    )
    repo = VehicleRepository(session)
    payloads: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=2):
        payloads.append(
            {
                "vehicle_number": row["vehicle_number"],
                "registration_number": row["registration_number"],
                "vendor_id": _parse_import_uuid(row.get("vendor_id"), field="vendor_id", row_number=index),
                "ward_id": _parse_import_uuid(row.get("ward_id"), field="ward_id", row_number=index),
                "route_id": _parse_import_uuid(row.get("route_id"), field="route_id", row_number=index, required=False),
                "vehicle_category": row.get("vehicle_category") or "primary",
                "secondary_waste_type": row.get("secondary_waste_type") or None,
                "truck_type": row.get("truck_type") or None,
                "capacity_kg": _parse_import_float(row.get("capacity_kg"), field="capacity_kg", row_number=index, default=0.0),
                "capacity_cubic_meter": _parse_import_float(
                    row.get("capacity_cubic_meter"), field="capacity_cubic_meter", row_number=index, default=0.0
                ),
                "fuel_type": row.get("fuel_type") or "diesel",
                "operational_status": row.get("operational_status") or "operational",
                "chassis_number": row.get("chassis_number") or None,
                "engine_number": row.get("engine_number") or None,
                "manufacture_year": _parse_import_int(
                    row.get("manufacture_year"), field="manufacture_year", row_number=index, default=None
                ),
                "active": _parse_bool(row.get("active"), default=True),
                "metadata_json": {},
            }
        )
    created = await repo.bulk_create(payloads)
    return {"created": len(created)}


@router.get("/secondary-waste-types")
async def list_secondary_waste_types(
    _: RoleContext = Depends(require_roles("admin", "ops", "viewer")),
) -> list[dict[str, str]]:
    return SECONDARY_WASTE_TYPES


def _gts_to_dict(row: GtsPointORM) -> dict[str, Any]:
    return {
        "id": str(row.id),
        "name": row.name,
        "latitude": row.latitude,
        "longitude": row.longitude,
        "address": row.address,
        "zone_id": str(row.zone_id) if row.zone_id else None,
        "ward_id": str(row.ward_id) if row.ward_id else None,
        "is_active": row.active,
        "active": row.active,
        "created_at": row.created_at.isoformat() if row.created_at else None,
        "updated_at": row.updated_at.isoformat() if row.updated_at else None,
    }


async def _ensure_unique_gts_name(session: AsyncSession, *, name: str, ward_id: UUID | None, exclude_id: UUID | None = None) -> None:
    stmt = select(GtsPointORM).where(func.lower(GtsPointORM.name) == name.strip().lower())
    if ward_id is None:
        stmt = stmt.where(GtsPointORM.ward_id.is_(None))
    else:
        stmt = stmt.where(GtsPointORM.ward_id == ward_id)
    if exclude_id is not None:
        stmt = stmt.where(GtsPointORM.id != exclude_id)
    exists = (await session.execute(stmt)).scalars().first()
    if exists is not None:
        raise HTTPException(status_code=409, detail="GTS name already exists for this ward")


@router.get("/gts")
async def list_gts(
    q: str | None = Query(default=None),
    ward_id: UUID | None = Query(default=None),
    active: bool | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=200),
    _: RoleContext = Depends(require_roles("admin", "ops", "viewer")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    stmt = select(GtsPointORM)
    count_stmt = select(func.count(GtsPointORM.id))
    filters = []
    if q:
        filters.append(GtsPointORM.name.ilike(f"%{q}%"))
    if ward_id is not None:
        filters.append(GtsPointORM.ward_id == ward_id)
    if active is not None:
        filters.append(GtsPointORM.active.is_(active))
    for condition in filters:
        stmt = stmt.where(condition)
        count_stmt = count_stmt.where(condition)
    total = int((await session.execute(count_stmt)).scalar() or 0)
    rows = (
        await session.execute(
            stmt.order_by(GtsPointORM.name.asc()).offset((page - 1) * page_size).limit(page_size)
        )
    ).scalars().all()
    return {"items": [_gts_to_dict(row) for row in rows], "total": total, "page": page, "page_size": page_size}


@router.post("/gts")
async def create_gts(
    payload: GtsIn,
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    await _ensure_unique_gts_name(session, name=payload.name, ward_id=payload.ward_id)
    row = GtsPointORM(
        name=payload.name.strip(),
        latitude=payload.latitude,
        longitude=payload.longitude,
        address=payload.address,
        zone_id=payload.zone_id,
        ward_id=payload.ward_id,
        active=payload.is_active,
    )
    session.add(row)
    await session.commit()
    await session.refresh(row)
    return _gts_to_dict(row)


@router.put("/gts/{gts_id}")
async def update_gts(
    gts_id: UUID,
    payload: GtsIn,
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    row = (await session.execute(select(GtsPointORM).where(GtsPointORM.id == gts_id))).scalars().first()
    if row is None:
        _raise_not_found("GTS", gts_id)
    await _ensure_unique_gts_name(session, name=payload.name, ward_id=payload.ward_id, exclude_id=gts_id)
    row.name = payload.name.strip()
    row.latitude = payload.latitude
    row.longitude = payload.longitude
    row.address = payload.address
    row.zone_id = payload.zone_id
    row.ward_id = payload.ward_id
    row.active = payload.is_active
    await session.commit()
    await session.refresh(row)
    return _gts_to_dict(row)


@router.delete("/gts/{gts_id}", response_model=MessageResponse)
async def delete_gts(
    gts_id: UUID,
    _: RoleContext = Depends(require_roles("admin")),
    session: AsyncSession = Depends(get_db_session),
) -> MessageResponse:
    row = (await session.execute(select(GtsPointORM).where(GtsPointORM.id == gts_id))).scalars().first()
    if row is None:
        _raise_not_found("GTS", gts_id)
    await session.delete(row)
    await session.commit()
    return MessageResponse(message="deleted")


@router.get("/dump-yards")
async def list_dump_yards(
    active: bool | None = Query(default=None),
    zone_id: UUID | None = Query(default=None),
    q: str | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=200, ge=1, le=200),
    _: RoleContext = Depends(require_roles("admin", "ops", "viewer")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    stmt = select(DumpYardORM).order_by(DumpYardORM.dump_yard_name.asc())
    count_stmt = select(func.count(DumpYardORM.id))
    if active is not None:
        stmt = stmt.where(DumpYardORM.active.is_(active))
        count_stmt = count_stmt.where(DumpYardORM.active.is_(active))
    if zone_id is not None:
        stmt = stmt.where(DumpYardORM.zone_id == zone_id)
        count_stmt = count_stmt.where(DumpYardORM.zone_id == zone_id)
    if q:
        stmt = stmt.where(DumpYardORM.dump_yard_name.ilike(f"%{q}%"))
        count_stmt = count_stmt.where(DumpYardORM.dump_yard_name.ilike(f"%{q}%"))
    total = int((await session.execute(count_stmt)).scalar() or 0)
    rows = (await session.execute(stmt.offset((page - 1) * page_size).limit(page_size))).scalars().all()
    return {"items": [_dump_yard_to_dict(row) for row in rows], "total": total, "page": page, "page_size": page_size}


def _dump_yard_payload(payload: DumpYardIn) -> dict[str, Any]:
    name = (payload.dump_yard_name or payload.name or "").strip()
    if not name:
        raise HTTPException(status_code=422, detail="Dump Yard name is required")
    code = (payload.dump_yard_code or name.upper().replace(" ", "-")[:32]).strip().upper()
    return {
        "dump_yard_code": code,
        "dump_yard_name": name,
        "address": payload.address,
        "capacity": payload.capacity,
        "zone_id": payload.zone_id,
        "ward_id": payload.ward_id,
        "lat": payload.lat if payload.lat is not None else payload.latitude,
        "lng": payload.lng if payload.lng is not None else payload.longitude,
        "active": payload.active if payload.is_active is None else payload.is_active,
    }


def _dump_yard_to_dict(row: DumpYardORM) -> dict[str, Any]:
    return {
        "id": str(row.id),
        "dump_yard_code": row.dump_yard_code,
        "dump_yard_name": row.dump_yard_name,
        "name": row.dump_yard_name,
        "address": row.address,
        "capacity": row.capacity,
        "zone_id": str(row.zone_id) if row.zone_id else None,
        "ward_id": str(row.ward_id) if row.ward_id else None,
        "lat": row.lat,
        "lng": row.lng,
        "latitude": row.lat,
        "longitude": row.lng,
        "active": row.active,
        "is_active": row.active,
        "created_at": row.created_at.isoformat() if row.created_at else None,
        "updated_at": row.updated_at.isoformat() if row.updated_at else None,
    }


async def _ensure_unique_dump_yard_name(session: AsyncSession, *, name: str, exclude_id: UUID | None = None) -> None:
    stmt = select(DumpYardORM).where(func.lower(DumpYardORM.dump_yard_name) == name.strip().lower())
    if exclude_id is not None:
        stmt = stmt.where(DumpYardORM.id != exclude_id)
    exists = (await session.execute(stmt)).scalars().first()
    if exists is not None:
        raise HTTPException(status_code=409, detail="Dump Yard name already exists")


@router.post("/dump-yards")
async def create_dump_yard(
    payload: DumpYardIn,
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    data = _dump_yard_payload(payload)
    await _ensure_unique_dump_yard_name(session, name=data["dump_yard_name"])
    row = DumpYardORM(**data)
    session.add(row)
    await session.commit()
    await session.refresh(row)
    return _dump_yard_to_dict(row)


@router.put("/dump-yards/{dump_yard_id}")
async def update_dump_yard(
    dump_yard_id: UUID,
    payload: DumpYardIn,
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    row = (await session.execute(select(DumpYardORM).where(DumpYardORM.id == dump_yard_id))).scalars().first()
    if row is None:
        _raise_not_found("dump yard", dump_yard_id)
    data = _dump_yard_payload(payload)
    await _ensure_unique_dump_yard_name(session, name=data["dump_yard_name"], exclude_id=dump_yard_id)
    for key, value in data.items():
        setattr(row, key, value)
    await session.commit()
    await session.refresh(row)
    return _dump_yard_to_dict(row)


@router.delete("/dump-yards/{dump_yard_id}", response_model=MessageResponse)
async def delete_dump_yard(
    dump_yard_id: UUID,
    _: RoleContext = Depends(require_roles("admin")),
    session: AsyncSession = Depends(get_db_session),
) -> MessageResponse:
    row = (await session.execute(select(DumpYardORM).where(DumpYardORM.id == dump_yard_id))).scalars().first()
    if row is None:
        _raise_not_found("dump yard", dump_yard_id)
    await session.delete(row)
    await session.commit()
    return MessageResponse(message="deleted")


def _assignment_to_dict(row: SecondaryVehicleAssignmentORM) -> dict[str, Any]:
    item = _to_dict(row)
    vehicle = getattr(row, "vehicle", None)
    gtc = getattr(row, "gtc_pickup_point", None)
    dump_yard = getattr(row, "dump_yard", None)
    item.update(
        {
            "vehicle_number": getattr(vehicle, "vehicle_number", None),
            "registration_number": getattr(vehicle, "registration_number", None),
            "secondary_waste_type": getattr(vehicle, "secondary_waste_type", None),
            "gtc_name": getattr(gtc, "pickup_name", None),
            "gts_name": getattr(gtc, "pickup_name", None),
            "dump_yard_name": getattr(dump_yard, "dump_yard_name", None),
        }
    )
    return item


@router.get("/secondary-vehicle-assignments")
async def list_secondary_vehicle_assignments(
    vehicle_id: UUID | None = Query(default=None),
    active: bool | None = Query(default=None),
    material_type: str | None = Query(default=None),
    _: RoleContext = Depends(require_roles("admin", "ops", "viewer")),
    session: AsyncSession = Depends(get_db_session),
) -> list[dict[str, Any]]:
    stmt = (
        select(SecondaryVehicleAssignmentORM)
        .where(True)
        .order_by(SecondaryVehicleAssignmentORM.created_at.desc())
    )
    if vehicle_id is not None:
        stmt = stmt.where(SecondaryVehicleAssignmentORM.vehicle_id == vehicle_id)
    if active is not None:
        stmt = stmt.where(SecondaryVehicleAssignmentORM.active.is_(active))
    if material_type:
        stmt = stmt.where(SecondaryVehicleAssignmentORM.material_type == material_type)
    rows = (await session.execute(stmt)).scalars().all()
    for row in rows:
        await session.refresh(row, attribute_names=["vehicle", "gtc_pickup_point", "dump_yard"])
    return [_assignment_to_dict(row) for row in rows]


@router.post("/secondary-vehicle-assignments")
async def create_secondary_vehicle_assignment(
    payload: SecondaryVehicleAssignmentIn,
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    vehicle = (await session.execute(select(VehicleORM).where(VehicleORM.id == payload.vehicle_id))).scalars().first()
    if vehicle is None:
        _raise_not_found("vehicle", payload.vehicle_id)
    if vehicle.vehicle_category != "secondary":
        raise HTTPException(status_code=422, detail="vehicle must be a secondary vehicle")
    row = SecondaryVehicleAssignmentORM(
        **payload.model_dump(exclude={"assigned_from"}),
        assigned_from=payload.assigned_from or datetime.now(timezone.utc),
    )
    session.add(row)
    await session.commit()
    await session.refresh(row)
    return _to_dict(row)


@router.put("/secondary-vehicle-assignments/{assignment_id}")
async def update_secondary_vehicle_assignment(
    assignment_id: UUID,
    payload: SecondaryVehicleAssignmentIn,
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    row = (
        await session.execute(select(SecondaryVehicleAssignmentORM).where(SecondaryVehicleAssignmentORM.id == assignment_id))
    ).scalars().first()
    if row is None:
        _raise_not_found("secondary vehicle assignment", assignment_id)
    vehicle = (await session.execute(select(VehicleORM).where(VehicleORM.id == payload.vehicle_id))).scalars().first()
    if vehicle is None:
        _raise_not_found("vehicle", payload.vehicle_id)
    if vehicle.vehicle_category != "secondary":
        raise HTTPException(status_code=422, detail="vehicle must be a secondary vehicle")
    for key, value in payload.model_dump().items():
        setattr(row, key, value)
    if row.assigned_from is None:
        row.assigned_from = datetime.now(timezone.utc)
    await session.commit()
    await session.refresh(row)
    return _to_dict(row)


@router.delete("/secondary-vehicle-assignments/{assignment_id}", response_model=MessageResponse)
async def delete_secondary_vehicle_assignment(
    assignment_id: UUID,
    _: RoleContext = Depends(require_roles("admin")),
    session: AsyncSession = Depends(get_db_session),
) -> MessageResponse:
    row = (
        await session.execute(select(SecondaryVehicleAssignmentORM).where(SecondaryVehicleAssignmentORM.id == assignment_id))
    ).scalars().first()
    if row is None:
        _raise_not_found("secondary vehicle assignment", assignment_id)
    await session.delete(row)
    await session.commit()
    return MessageResponse(message="deleted")


def _weighment_to_dict(row: DumpYardWeighmentORM) -> dict[str, Any]:
    item = _to_dict(row)
    item["service_date"] = row.service_date.isoformat() if row.service_date else None
    vehicle = getattr(row, "vehicle", None)
    gtc = getattr(row, "gtc_pickup_point", None)
    dump_yard = getattr(row, "dump_yard", None)
    ward = getattr(vehicle, "ward", None)
    route = getattr(vehicle, "route", None)
    zone = getattr(ward, "zone", None)
    item.update(
        {
            "vehicle_number": getattr(vehicle, "vehicle_number", None),
            "registration_number": getattr(vehicle, "registration_number", None),
            "vehicle_category": getattr(vehicle, "vehicle_category", None),
            "vehicle_type": getattr(vehicle, "truck_type", None),
            "route_id": str(getattr(route, "id", "")) if route else None,
            "route_name": getattr(route, "route_name", None),
            "ward_id": str(getattr(ward, "id", "")) if ward else None,
            "ward_name": getattr(ward, "ward_name", None),
            "zone_id": str(getattr(zone, "id", "")) if zone else None,
            "zone_name": getattr(zone, "zone_name", None),
            "gtc_name": getattr(gtc, "pickup_name", None),
            "gts_name": getattr(gtc, "pickup_name", None),
            "gts_pickup_point_id": str(getattr(gtc, "id", "")) if gtc else None,
            "material_label": _material_label(row.material_type),
            "dump_yard_name": getattr(dump_yard, "dump_yard_name", None),
        }
    )
    return item


async def _fetch_dump_yard_weighment_rows(
    session: AsyncSession,
    *,
    date_from: date | None = None,
    date_to: date | None = None,
    vehicle_id: UUID | None = None,
    material_type: str | None = None,
    dump_yard_id: UUID | None = None,
) -> list[DumpYardWeighmentORM]:
    stmt = select(DumpYardWeighmentORM).order_by(DumpYardWeighmentORM.entry_time.desc())
    if date_from is not None:
        stmt = stmt.where(DumpYardWeighmentORM.service_date >= date_from)
    if date_to is not None:
        stmt = stmt.where(DumpYardWeighmentORM.service_date <= date_to)
    if vehicle_id is not None:
        stmt = stmt.where(DumpYardWeighmentORM.vehicle_id == vehicle_id)
    if material_type:
        stmt = stmt.where(DumpYardWeighmentORM.material_type == material_type)
    if dump_yard_id is not None:
        stmt = stmt.where(DumpYardWeighmentORM.dump_yard_id == dump_yard_id)
    rows = (await session.execute(stmt)).scalars().all()
    for row in rows:
        await session.refresh(row, attribute_names=["vehicle", "gtc_pickup_point", "dump_yard"])
        if row.vehicle:
            await session.refresh(row.vehicle, attribute_names=["ward", "route"])
            if row.vehicle.ward:
                await session.refresh(row.vehicle.ward, attribute_names=["zone"])
    return rows


@router.get("/dump-yard-weighment")
@router.get("/dump-yard-weighments")
async def list_dump_yard_weighments(
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    vehicle_id: UUID | None = Query(default=None),
    material_type: str | None = Query(default=None),
    dump_yard_id: UUID | None = Query(default=None),
    _: RoleContext = Depends(require_roles("admin", "ops", "viewer")),
    session: AsyncSession = Depends(get_db_session),
) -> list[dict[str, Any]]:
    rows = await _fetch_dump_yard_weighment_rows(
        session,
        date_from=date_from,
        date_to=date_to,
        vehicle_id=vehicle_id,
        material_type=material_type,
        dump_yard_id=dump_yard_id,
    )
    return [_weighment_to_dict(row) for row in rows]


@router.get("/dump-yard-weighment/export.xlsx")
async def export_dump_yard_weighments_xlsx(
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    vehicle_id: UUID | None = Query(default=None),
    material_type: str | None = Query(default=None),
    dump_yard_id: UUID | None = Query(default=None),
    _: RoleContext = Depends(require_roles("admin", "ops", "viewer")),
    session: AsyncSession = Depends(get_db_session),
) -> StreamingResponse:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    rows = await _fetch_dump_yard_weighment_rows(
        session,
        date_from=date_from,
        date_to=date_to,
        vehicle_id=vehicle_id,
        material_type=material_type,
        dump_yard_id=dump_yard_id,
    )
    records = [_weighment_to_dict(row) for row in rows]
    period_from = date_from or (min((row.service_date for row in rows), default=datetime.now(timezone.utc).date()))
    period_to = date_to or (max((row.service_date for row in rows), default=datetime.now(timezone.utc).date()))

    wb = Workbook()
    ws = wb.active
    ws.title = "Executive Summary"
    detail_ws = wb.create_sheet("Detail Records")
    daily_ws = wb.create_sheet("Daily Trend")
    material_ws = wb.create_sheet("Material Summary")
    vehicle_ws = wb.create_sheet("Vehicle Summary")

    dark = "0F766E"
    dark2 = "064E3B"
    green = "D1FAE5"
    pale = "F8FAFC"
    amber = "FFFBEB"
    line = Side(style="thin", color="CBD5E1")
    border = Border(bottom=line)

    def title(sheet, text: str, subtitle: str, cols: int) -> None:
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
        sheet["A1"] = text
        sheet["A2"] = subtitle
        sheet["A1"].font = Font(name="Aptos Display", size=20, bold=True, color=dark2)
        sheet["A1"].fill = PatternFill("solid", fgColor=green)
        sheet["A1"].alignment = Alignment(horizontal="center")
        sheet["A2"].font = Font(name="Aptos", size=10, color="475569")
        sheet["A2"].fill = PatternFill("solid", fgColor=pale)
        sheet["A2"].alignment = Alignment(horizontal="center")

    def header(sheet, row_idx: int, values: list[str]) -> None:
        for col_idx, value in enumerate(values, start=1):
            cell = sheet.cell(row_idx, col_idx, value)
            cell.font = Font(name="Aptos", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=dark)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    def style_table(sheet, min_row: int, max_row: int, max_col: int) -> None:
        for row_idx in range(min_row, max_row + 1):
            fill = PatternFill("solid", fgColor="FFFFFF" if row_idx % 2 else pale)
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row_idx, col_idx)
                cell.fill = fill
                cell.border = border
                cell.font = Font(name="Aptos", size=10, color="0F172A", bold=col_idx == 1)
                if isinstance(cell.value, int | float):
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")

    material_totals: dict[str, dict[str, float]] = {}
    daily_totals: dict[str, dict[str, float]] = {}
    vehicle_totals: dict[str, dict[str, float]] = {}
    total_net = total_gross = total_tare = 0.0
    for record in records:
        material = record.get("material_label") or str(record.get("material_type") or "-").replace("_", " ").upper()
        day = record.get("service_date") or "-"
        vehicle = record.get("vehicle_number") or record.get("registration_number") or "-"
        gross = float(record.get("gross_weight_kg") or 0)
        tare = float(record.get("tare_weight_kg") or 0)
        net = float(record.get("net_weight_kg") or 0)
        total_gross += gross
        total_tare += tare
        total_net += net
        for bucket, key in ((material_totals, material), (daily_totals, day), (vehicle_totals, vehicle)):
            item = bucket.setdefault(key, {"entries": 0, "gross": 0.0, "tare": 0.0, "net": 0.0})
            item["entries"] += 1
            item["gross"] += gross
            item["tare"] += tare
            item["net"] += net

    title(ws, "Dump Yard Weighment Report", f"Period: {period_from} to {period_to} | Generated: {datetime.now().strftime('%d %b %Y %H:%M')}", 8)
    kpis = [
        ("Total Entries", len(records), "records"),
        ("Net Collection", total_net / 1000, "tons"),
        ("Gross Weight", total_gross / 1000, "tons"),
        ("Tare Weight", total_tare / 1000, "tons"),
        ("Average Net / Entry", total_net / max(len(records), 1), "kg"),
        ("Active Trucks", len(vehicle_totals), "trucks"),
        ("Material Types", len(material_totals), "types"),
    ]
    header(ws, 4, ["KPI", "Value", "Unit", "Operational Reading"])
    for idx, (name, value, unit) in enumerate(kpis, start=5):
        ws.cell(idx, 1, name)
        ws.cell(idx, 2, value)
        ws.cell(idx, 3, unit)
        ws.cell(idx, 4, "Use this workbook for reconciliation, material analytics, and route-wise tonnage review." if idx == 5 else "")
    style_table(ws, 5, 5 + len(kpis) - 1, 4)
    ws.cell(14, 1, "Professional workbook tabs: Daily Trend, Material Summary, Vehicle Summary, and Detail Records.")
    ws.cell(14, 1).fill = PatternFill("solid", fgColor=amber)
    ws.cell(14, 1).font = Font(bold=True, color=dark2)

    def write_summary(sheet, sheet_title: str, data: dict[str, dict[str, float]], first_col: str, chart_kind: str) -> None:
        title(sheet, sheet_title, f"Period: {period_from} to {period_to}", 6)
        header(sheet, 4, [first_col, "Entries", "Gross KG", "Tare KG", "Net KG", "Net Ton"])
        ordered = sorted(data.items(), key=lambda item: item[1]["net"], reverse=True)
        if first_col == "Date":
            ordered = sorted(data.items(), key=lambda item: item[0])
        for row_idx, (key, item) in enumerate(ordered, start=5):
            sheet.append([key, int(item["entries"]), item["gross"], item["tare"], item["net"], item["net"] / 1000])
        if ordered:
            style_table(sheet, 5, 5 + len(ordered) - 1, 6)
            chart = LineChart() if chart_kind == "line" else BarChart()
            chart.title = sheet_title
            chart.y_axis.title = "Net KG"
            chart.x_axis.title = first_col
            data_ref = Reference(sheet, min_col=5, min_row=4, max_row=4 + len(ordered))
            cats = Reference(sheet, min_col=1, min_row=5, max_row=4 + len(ordered))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 8
            chart.width = 18
            sheet.add_chart(chart, "H4")
        sheet.freeze_panes = "A5"

    write_summary(daily_ws, "Seven Day Collection Trend", daily_totals, "Date", "line")
    write_summary(material_ws, "Material Wise Collection", material_totals, "Material", "bar")
    write_summary(vehicle_ws, "Vehicle Wise Collection", vehicle_totals, "Vehicle", "bar")

    detail_headers = [
        "Date", "Entry Time", "Vehicle", "Registration", "Zone", "Ward", "Route", "GTS", "Dump Yard",
        "Material", "Gross KG", "Tare KG", "Net KG", "Net Ton", "Slip Number", "Operator", "Remarks",
    ]
    title(detail_ws, "Detailed Weighment Register", f"Period: {period_from} to {period_to}", len(detail_headers))
    header(detail_ws, 4, detail_headers)
    for record in records:
        detail_ws.append(
            [
                record.get("service_date"),
                record.get("entry_time"),
                record.get("vehicle_number"),
                record.get("registration_number"),
                record.get("zone_name"),
                record.get("ward_name"),
                record.get("route_name"),
                record.get("gts_name"),
                record.get("dump_yard_name"),
                record.get("material_label") or str(record.get("material_type") or "").replace("_", " ").upper(),
                float(record.get("gross_weight_kg") or 0),
                float(record.get("tare_weight_kg") or 0),
                float(record.get("net_weight_kg") or 0),
                float(record.get("net_weight_kg") or 0) / 1000,
                record.get("slip_number"),
                record.get("operator_name"),
                record.get("remarks"),
            ]
        )
    if records:
        style_table(detail_ws, 5, 5 + len(records) - 1, len(detail_headers))
    detail_ws.freeze_panes = "A5"

    for sheet in wb.worksheets:
        for col_idx in range(1, sheet.max_column + 1):
            max_len = max((len(str(sheet.cell(row_idx, col_idx).value or "")) for row_idx in range(1, sheet.max_row + 1)), default=12)
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 3, 12), 36)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"dump_yard_weighment_{period_from}_{period_to}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


async def _ensure_weighment_can_save(
    session: AsyncSession,
    payload: DumpYardWeighmentIn,
    *,
    entry_time: datetime,
    service_date: date,
    exclude_id: UUID | None = None,
) -> None:
    if payload.gross_weight_kg < payload.tare_weight_kg:
        raise HTTPException(status_code=422, detail="gross_weight_kg cannot be less than tare_weight_kg")
    if payload.net_weight_kg is not None and payload.net_weight_kg < 0:
        raise HTTPException(status_code=422, detail="net_weight_kg cannot be negative")
    duplicate_start = entry_time - timedelta(minutes=2)
    duplicate_end = entry_time + timedelta(minutes=2)
    duplicate_stmt = select(DumpYardWeighmentORM).where(
        DumpYardWeighmentORM.vehicle_id == payload.vehicle_id,
        DumpYardWeighmentORM.dump_yard_id == payload.dump_yard_id,
        DumpYardWeighmentORM.material_type == payload.material_type,
        DumpYardWeighmentORM.service_date == service_date,
        DumpYardWeighmentORM.entry_time >= duplicate_start,
        DumpYardWeighmentORM.entry_time <= duplicate_end,
    )
    if exclude_id is not None:
        duplicate_stmt = duplicate_stmt.where(DumpYardWeighmentORM.id != exclude_id)
    if payload.slip_number:
        slip_stmt = select(DumpYardWeighmentORM).where(DumpYardWeighmentORM.slip_number == payload.slip_number)
        if exclude_id is not None:
            slip_stmt = slip_stmt.where(DumpYardWeighmentORM.id != exclude_id)
        if (await session.execute(slip_stmt)).scalars().first() is not None:
            raise HTTPException(status_code=409, detail="weighment slip number already exists")
    if (await session.execute(duplicate_stmt)).scalars().first() is not None:
        raise HTTPException(status_code=409, detail="possible duplicate weighment entry within 2 minutes")


@router.post("/dump-yard-weighment")
@router.post("/dump-yard-weighments")
async def create_dump_yard_weighment(
    payload: DumpYardWeighmentIn,
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    entry_time = payload.entry_time or datetime.now(timezone.utc)
    service_date = payload.service_date or entry_time.date()
    await _ensure_weighment_can_save(session, payload, entry_time=entry_time, service_date=service_date)
    net_weight = payload.net_weight_kg
    if net_weight is None:
        net_weight = float(payload.gross_weight_kg) - float(payload.tare_weight_kg)
    row = DumpYardWeighmentORM(
        assignment_id=payload.assignment_id,
        vehicle_id=payload.vehicle_id,
        gtc_pickup_point_id=payload.gtc_pickup_point_id,
        dump_yard_id=payload.dump_yard_id,
        material_type=payload.material_type,
        service_date=service_date,
        entry_time=entry_time,
        gross_weight_kg=payload.gross_weight_kg,
        tare_weight_kg=payload.tare_weight_kg,
        net_weight_kg=net_weight,
        slip_number=payload.slip_number,
        operator_name=payload.operator_name,
        remarks=payload.remarks,
    )
    session.add(row)
    await session.commit()
    await session.refresh(row)
    await session.refresh(row, attribute_names=["vehicle", "gtc_pickup_point", "dump_yard"])
    if row.vehicle:
        await session.refresh(row.vehicle, attribute_names=["ward", "route"])
        if row.vehicle.ward:
            await session.refresh(row.vehicle.ward, attribute_names=["zone"])
    return _weighment_to_dict(row)


@router.put("/dump-yard-weighment/{weighment_id}")
@router.put("/dump-yard-weighments/{weighment_id}")
async def update_dump_yard_weighment(
    weighment_id: UUID,
    payload: DumpYardWeighmentIn,
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    row = (
        await session.execute(select(DumpYardWeighmentORM).where(DumpYardWeighmentORM.id == weighment_id))
    ).scalars().first()
    if row is None:
        _raise_not_found("dump yard weighment", weighment_id)
    entry_time = payload.entry_time or row.entry_time or datetime.now(timezone.utc)
    service_date = payload.service_date or entry_time.date()
    await _ensure_weighment_can_save(
        session,
        payload,
        entry_time=entry_time,
        service_date=service_date,
        exclude_id=weighment_id,
    )
    net_weight = payload.net_weight_kg
    if net_weight is None:
        net_weight = float(payload.gross_weight_kg) - float(payload.tare_weight_kg)
    row.assignment_id = payload.assignment_id
    row.vehicle_id = payload.vehicle_id
    row.gtc_pickup_point_id = payload.gtc_pickup_point_id
    row.dump_yard_id = payload.dump_yard_id
    row.material_type = payload.material_type
    row.service_date = service_date
    row.entry_time = entry_time
    row.gross_weight_kg = payload.gross_weight_kg
    row.tare_weight_kg = payload.tare_weight_kg
    row.net_weight_kg = net_weight
    row.slip_number = payload.slip_number
    row.operator_name = payload.operator_name
    row.remarks = payload.remarks
    await session.commit()
    await session.refresh(row)
    await session.refresh(row, attribute_names=["vehicle", "gtc_pickup_point", "dump_yard"])
    if row.vehicle:
        await session.refresh(row.vehicle, attribute_names=["ward", "route"])
        if row.vehicle.ward:
            await session.refresh(row.vehicle.ward, attribute_names=["zone"])
    return _weighment_to_dict(row)


@router.get("/routes")
async def list_routes(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=200),
    q: str | None = Query(default=None),
    active: bool | None = Query(default=None),
    ward_id: UUID | None = Query(default=None),
    zone_id: UUID | None = Query(default=None),
    ui_compat: bool = Query(default=False),
    sort_by: str = Query(default="created_at"),
    sort_order: str = Query(default="desc", pattern="^(asc|desc)$"),
    _: RoleContext = Depends(require_roles("admin", "ops", "viewer")),
    session: AsyncSession = Depends(get_db_session),
) -> PageResponse | list[dict[str, Any]]:
    result = await _list_entities(
        session,
        RouteORM,
        page=page,
        page_size=page_size,
        q=q,
        sort_by=sort_by,
        sort_order=sort_order,
        filters={},
    )

    if not ui_compat and ward_id is None and zone_id is None:
        return result

    compat_items: list[dict[str, Any]] = []
    route_ids = [UUID(str(route.get("id"))) for route in result.items if route.get("id")]
    points_by_route: dict[str, list[PickupPointORM]] = {str(route_id): [] for route_id in route_ids}
    if route_ids:
        point_rows = (
            await session.execute(
                select(PickupPointORM)
                .where(PickupPointORM.route_id.in_(route_ids))
                .order_by(PickupPointORM.route_id.asc(), PickupPointORM.sequence_no.asc())
            )
        ).scalars().all()
        for point in point_rows:
            if point.route_id is not None:
                points_by_route.setdefault(str(point.route_id), []).append(point)

    for route in result.items:
        route_zone_id = str(route.get("zone_id") or "")
        route_ward_id = str(route.get("ward_id") or "")
        route_id = str(route.get("id") or "")
        stats = _route_stats_from_points(points_by_route.get(route_id, []))
        mapped = {
            "id": route_id,
            "name": route.get("route_name") or "",
            "code": route.get("route_name") or "",
            "route_name": route.get("route_name") or "",
            "type": route.get("route_type") or "primary",
            "route_type": route.get("route_type") or "primary",
            "zone_id": route_zone_id,
            "zoneId": route_zone_id,
            "ward_id": route_ward_id,
            "wardId": route_ward_id,
            "polyline_coordinates": route.get("polyline_coordinates") or [],
            "status": "active",
            "active": True,
            **stats,
        }

        if ward_id is not None and mapped["ward_id"] != str(ward_id):
            continue
        if zone_id is not None and mapped["zone_id"] != str(zone_id):
            continue

        compat_items.append(mapped)

    if ui_compat:
        return compat_items

    # Non-compat mode with zone/ward filters keeps existing paged envelope while applying filters.
    return PageResponse(items=compat_items, page=page, page_size=page_size, total=len(compat_items))


@router.post("/routes")
async def create_route(
    payload: RouteIn,
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    repo = RouteRepository(session)
    row = await repo.create(
        route_name=payload.route_name,
        route_type=payload.route_type,
        zone_id=payload.zone_id,
        ward_id=payload.ward_id,
        polyline_coordinates=payload.polyline_coordinates,
    )
    return _to_dict(row)


@router.get("/routes/{route_id}")
async def get_route(
    route_id: UUID,
    _: RoleContext = Depends(require_roles("admin", "ops", "viewer")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    repo = RouteRepository(session)
    row = await _fetch_or_404(repo.get_by_id, "route", route_id)
    return _to_dict(row)


@router.put("/routes/{route_id}")
async def update_route(
    route_id: UUID,
    payload: RouteIn,
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    repo = RouteRepository(session)
    try:
        row = await repo.update(
            route_id,
            route_name=payload.route_name,
            route_type=payload.route_type,
            zone_id=payload.zone_id,
            ward_id=payload.ward_id,
            polyline_coordinates=payload.polyline_coordinates,
        )
    except NoResultFound:
        _raise_not_found("route", route_id)
    return _to_dict(row)


@router.delete("/routes/{route_id}", response_model=MessageResponse)
async def delete_route(
    route_id: UUID,
    _: RoleContext = Depends(require_roles("admin")),
    session: AsyncSession = Depends(get_db_session),
) -> MessageResponse:
    repo = RouteRepository(session)
    try:
        await repo.delete(route_id)
    except NoResultFound:
        _raise_not_found("route", route_id)
    return MessageResponse(message="deleted")


@router.post("/routes/import")
async def bulk_import_routes(
    file: UploadFile = File(...),
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    content = (await file.read()).decode("utf-8")
    rows = _parse_csv_with_required(content, required_columns={"route_name", "zone_id", "ward_id", "polyline_coordinates"})
    repo = RouteRepository(session)
    payloads: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=2):
        payloads.append(
            {
                "route_name": row["route_name"],
                "zone_id": _parse_import_uuid(row.get("zone_id"), field="zone_id", row_number=index),
                "ward_id": _parse_import_uuid(row.get("ward_id"), field="ward_id", row_number=index),
                "polyline_coordinates": json.loads(row.get("polyline_coordinates") or "[]"),
            }
        )
    created = await repo.bulk_create(payloads)
    return {"created": len(created)}


@router.get("/geofences", response_model=PageResponse)
async def list_geofences(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=200),
    q: str | None = Query(default=None),
    active: bool | None = Query(default=None),
    geofence_for: str | None = Query(default=None),
    ward_id: UUID | None = Query(default=None),
    zone_id: UUID | None = Query(default=None),
    route_id: UUID | None = Query(default=None),
    type: str | None = Query(default=None),  # noqa: A002
    geometry_type: str | None = Query(default=None),
    sort_by: str = Query(default="created_at"),
    sort_order: str = Query(default="desc", pattern="^(asc|desc)$"),
    _: RoleContext = Depends(require_roles("admin", "ops", "viewer")),
    session: AsyncSession = Depends(get_db_session),
) -> PageResponse:
    filters: dict[str, Any] = {
        "active": active,
        "type": type,
        "geometry_type": geometry_type,
        "geofence_for": geofence_for,
        "zone_id": zone_id,
        "ward_id": ward_id,
        "route_id": route_id,
    }

    return await _list_entities(
        session,
        GeofenceORM,
        page=page,
        page_size=page_size,
        q=q,
        sort_by=sort_by,
        sort_order=sort_order,
        filters=filters,
    )


@router.post("/geofences")
async def create_geofence(
    payload: GeofenceIn,
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    repo = GeofenceRepository(session)
    if payload.geofence_for in {"ward", "route"} and payload.ward_id is None:
        raise HTTPException(status_code=422, detail="ward_id is required for ward/route geofence")
    if payload.geofence_for == "route" and payload.route_id is None:
        raise HTTPException(status_code=422, detail="route_id is required for route geofence")

    create_kwargs: dict[str, Any] = {
        "geofence_code": payload.geofence_code,
        "geofence_name": payload.geofence_name,
        "type": payload.type,
        "geometry_type": payload.geometry_type,
        "center_lat": payload.center_lat,
        "center_lng": payload.center_lng,
        "radius_meter": payload.radius_meter,
        "geofence_for": payload.geofence_for,
        "zone_id": payload.zone_id,
        "ward_id": payload.ward_id,
        "route_id": payload.route_id,
        "scope_type": "zone" if payload.geofence_for == "zone" else "ward",
        "scope_id": payload.zone_id if payload.geofence_for == "zone" else payload.ward_id,
        "active": payload.active,
    }
    if payload.polygon is not None:
        create_kwargs["polygon"] = payload.polygon
    row = await repo.create(**create_kwargs)
    return _to_dict(row)


@router.get("/geofences/{geofence_id}")
async def get_geofence(
    geofence_id: UUID,
    _: RoleContext = Depends(require_roles("admin", "ops", "viewer")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    repo = GeofenceRepository(session)
    row = await _fetch_or_404(repo.get_by_id, "geofence", geofence_id)
    return _to_dict(row)


@router.put("/geofences/{geofence_id}")
async def update_geofence(
    geofence_id: UUID,
    payload: GeofenceIn,
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    if payload.geofence_for in {"ward", "route"} and payload.ward_id is None:
        raise HTTPException(status_code=422, detail="ward_id is required for ward/route geofence")
    if payload.geofence_for == "route" and payload.route_id is None:
        raise HTTPException(status_code=422, detail="route_id is required for route geofence")

    repo = GeofenceRepository(session)
    try:
        update_kwargs: dict[str, Any] = {
            "geofence_code": payload.geofence_code,
            "geofence_name": payload.geofence_name,
            "type": payload.type,
            "geometry_type": payload.geometry_type,
            "center_lat": payload.center_lat,
            "center_lng": payload.center_lng,
            "radius_meter": payload.radius_meter,
            "geofence_for": payload.geofence_for,
            "zone_id": payload.zone_id,
            "ward_id": payload.ward_id,
            "route_id": payload.route_id,
            "scope_type": "zone" if payload.geofence_for == "zone" else "ward",
            "scope_id": payload.zone_id if payload.geofence_for == "zone" else payload.ward_id,
            "active": payload.active,
        }
        if payload.polygon is not None:
            update_kwargs["polygon"] = payload.polygon
        row = await repo.update(geofence_id, **update_kwargs)
    except NoResultFound:
        _raise_not_found("geofence", geofence_id)
    return _to_dict(row)


@router.delete("/geofences/{geofence_id}", response_model=MessageResponse)
async def delete_geofence(
    geofence_id: UUID,
    _: RoleContext = Depends(require_roles("admin")),
    session: AsyncSession = Depends(get_db_session),
) -> MessageResponse:
    repo = GeofenceRepository(session)
    try:
        await repo.delete(geofence_id)
    except NoResultFound:
        _raise_not_found("geofence", geofence_id)
    return MessageResponse(message="deleted")


@router.post("/geofences/import")
async def bulk_import_geofences(
    file: UploadFile = File(...),
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    content = (await file.read()).decode("utf-8")
    rows = _parse_csv_with_required(
        content,
        required_columns={"geofence_code", "geofence_name", "type", "geometry_type", "geofence_for", "zone_id", "coordinates"},
    )
    repo = GeofenceRepository(session)
    payloads: list[dict[str, Any]] = []

    def _parse_coordinates_polygon(raw: str, *, row_number: int) -> dict[str, Any]:
        chunks = [part.strip() for part in raw.split(";") if part.strip()]
        points: list[list[float]] = []
        for chunk in chunks:
            pair = [token.strip() for token in chunk.split(",")]
            if len(pair) != 2:
                raise HTTPException(status_code=400, detail=f"invalid coordinates at row {row_number}")
            try:
                lat = float(pair[0])
                lng = float(pair[1])
            except ValueError as exc:
                raise HTTPException(status_code=400, detail=f"invalid coordinates at row {row_number}") from exc
            points.append([lng, lat])
        if len(points) < 3:
            raise HTTPException(status_code=400, detail=f"coordinates must include at least 3 points at row {row_number}")
        if points[0] != points[-1]:
            points.append(points[0])
        return {"type": "Polygon", "coordinates": [points]}

    for index, row in enumerate(rows, start=2):
        geofence_for = (row.get("geofence_for") or "").strip().lower()
        if geofence_for not in {"zone", "ward", "route"}:
            raise HTTPException(status_code=400, detail=f"invalid geofence_for at row {index}")
        zone_id = _parse_import_uuid(row.get("zone_id"), field="zone_id", row_number=index, required=True)
        ward_id = _parse_import_uuid(row.get("ward_id"), field="ward_id", row_number=index, required=geofence_for != "zone")
        route_id = _parse_import_uuid(row.get("route_id"), field="route_id", row_number=index, required=geofence_for == "route")
        coordinates_raw = (row.get("coordinates") or "").strip()
        payloads.append(
            {
                "geofence_code": row["geofence_code"],
                "geofence_name": row["geofence_name"],
                "type": row["type"],
                "geometry_type": row["geometry_type"],
                "center_lat": _parse_import_float(row.get("center_lat"), field="center_lat", row_number=index, default=None),
                "center_lng": _parse_import_float(row.get("center_lng"), field="center_lng", row_number=index, default=None),
                "radius_meter": _parse_import_float(
                    row.get("radius_meter"), field="radius_meter", row_number=index, default=None
                ),
                "geofence_for": geofence_for,
                "zone_id": zone_id,
                "ward_id": ward_id,
                "route_id": route_id,
                "scope_type": "zone" if geofence_for == "zone" else "ward",
                "scope_id": zone_id if geofence_for == "zone" else ward_id,
                "polygon": _parse_coordinates_polygon(coordinates_raw, row_number=index) if coordinates_raw else None,
                "active": _parse_bool(row.get("active"), default=True),
            }
        )
    created = await repo.bulk_create(payloads)
    return {"created": len(created)}


@router.get("/wards", response_model=PageResponse)
async def list_wards(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=200),
    q: str | None = Query(default=None),
    active: bool | None = Query(default=None),
    zone_name: str | None = Query(default=None),
    sort_by: str = Query(default="created_at"),
    sort_order: str = Query(default="desc", pattern="^(asc|desc)$"),
    _: RoleContext = Depends(require_roles("admin", "ops", "viewer")),
    session: AsyncSession = Depends(get_db_session),
) -> PageResponse:
    return await _list_entities(
        session,
        WardORM,
        page=page,
        page_size=page_size,
        q=q,
        sort_by=sort_by,
        sort_order=sort_order,
        filters={"active": active, "zone_name": zone_name},
    )


@router.post("/wards")
async def create_ward(
    payload: WardIn,
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    repo = WardRepository(session)
    zone_id = await _resolve_zone_id(session, payload.zone_name)
    row = await repo.create(
        ward_code=payload.ward_code,
        ward_name=payload.ward_name,
        zone_id=zone_id,
        active=payload.active,
        population=payload.population,
        area=payload.area,
        total_pickup_points=payload.total_pickup_points,
    )
    return _to_dict(row)


@router.get("/wards/{ward_id}")
async def get_ward(
    ward_id: UUID,
    _: RoleContext = Depends(require_roles("admin", "ops", "viewer")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    repo = WardRepository(session)
    row = await _fetch_or_404(repo.get_by_id, "ward", ward_id)
    return _to_dict(row)


@router.put("/wards/{ward_id}")
async def update_ward(
    ward_id: UUID,
    payload: WardIn,
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    repo = WardRepository(session)
    zone_id = await _resolve_zone_id(session, payload.zone_name)
    try:
        row = await repo.update(
            ward_id,
            ward_code=payload.ward_code,
            ward_name=payload.ward_name,
            zone_id=zone_id,
            active=payload.active,
            population=payload.population,
            area=payload.area,
            total_pickup_points=payload.total_pickup_points,
        )
    except NoResultFound:
        _raise_not_found("ward", ward_id)
    return _to_dict(row)


@router.delete("/wards/{ward_id}", response_model=MessageResponse)
async def delete_ward(
    ward_id: UUID,
    _: RoleContext = Depends(require_roles("admin")),
    session: AsyncSession = Depends(get_db_session),
) -> MessageResponse:
    repo = WardRepository(session)
    try:
        await repo.delete(ward_id)
    except NoResultFound:
        _raise_not_found("ward", ward_id)
    return MessageResponse(message="deleted")


@router.post("/wards/import")
async def bulk_import_wards(
    file: UploadFile = File(...),
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    content = (await file.read()).decode("utf-8")
    rows = _parse_csv_with_required(content, required_columns={"ward_code", "ward_name", "zone_name"})
    repo = WardRepository(session)
    payloads: list[dict[str, Any]] = []
    for row in rows:
        zone_id = await _resolve_zone_id(session, row["zone_name"])
        payloads.append(
            {
                "ward_code": row["ward_code"],
                "ward_name": row["ward_name"],
                "zone_id": zone_id,
                "active": _parse_bool(row.get("active"), default=True),
            }
        )
    created = await repo.bulk_create(payloads)
    return {"created": len(created)}


@router.post("/device-assignments")
async def assign_device(
    payload: DeviceAssignmentIn,
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    svc = DeviceVehicleAssignmentService(DeviceVehicleAssignmentRepository(session))
    row = await svc.assign(
        AssignmentCreateInput(
            device_id=payload.device_id,
            vehicle_id=payload.vehicle_id,
            assigned_from=payload.assigned_from,
            remarks=payload.remarks,
        )
    )
    return _to_dict(row)


@router.get("/device-assignments/{device_id}")
async def get_active_assignment_by_device(
    device_id: UUID,
    _: RoleContext = Depends(require_roles("admin", "ops", "viewer")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    repo = DeviceVehicleAssignmentRepository(session)
    row = await repo.get_active_by_device(device_id)
    if row is None:
        _raise_not_found("device assignment", device_id)
    return _to_dict(row)


@router.put("/device-assignments/{device_id}")
async def reassign_device(
    device_id: UUID,
    vehicle_id: UUID = Query(...),
    remarks: str | None = Query(default=None),
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    svc = DeviceVehicleAssignmentService(DeviceVehicleAssignmentRepository(session))
    row = await svc.assign(
        AssignmentCreateInput(
            device_id=device_id,
            vehicle_id=vehicle_id,
            remarks=remarks,
        )
    )
    return _to_dict(row)


@router.delete("/device-assignments/{device_id}", response_model=MessageResponse)
async def unassign_device(
    device_id: UUID,
    remarks: str | None = Query(default=None),
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> MessageResponse:
    svc = DeviceVehicleAssignmentService(DeviceVehicleAssignmentRepository(session))
    row = await svc.unassign_device(device_id, remarks=remarks)
    if row is None:
        _raise_not_found("device assignment", device_id)
    return MessageResponse(message="deleted")


@router.get("/device-assignments", response_model=PageResponse)
async def list_device_assignments(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=200),
    device_id: UUID | None = Query(default=None),
    vehicle_id: UUID | None = Query(default=None),
    active: bool | None = Query(default=None),
    sort_by: str = Query(default="assigned_from"),
    sort_order: str = Query(default="desc", pattern="^(asc|desc)$"),
    _: RoleContext = Depends(require_roles("admin", "ops", "viewer")),
    session: AsyncSession = Depends(get_db_session),
) -> PageResponse:
    return await _list_entities(
        session,
        DeviceVehicleAssignmentORM,
        page=page,
        page_size=page_size,
        q=None,
        sort_by=sort_by,
        sort_order=sort_order,
        filters={"device_id": device_id, "vehicle_id": vehicle_id, "active": active},
    )


@router.post("/device-assignments/import")
async def bulk_import_device_assignments(
    file: UploadFile = File(...),
    _: RoleContext = Depends(require_roles("admin", "ops")),
    session: AsyncSession = Depends(get_db_session),
) -> dict[str, Any]:
    content = (await file.read()).decode("utf-8")
    rows = _parse_csv_with_required(content, required_columns={"device_id", "vehicle_id"})
    svc = DeviceVehicleAssignmentService(DeviceVehicleAssignmentRepository(session))
    created = 0
    for index, row in enumerate(rows, start=2):
        await svc.assign(
            AssignmentCreateInput(
                device_id=_parse_import_uuid(row.get("device_id"), field="device_id", row_number=index),
                vehicle_id=_parse_import_uuid(row.get("vehicle_id"), field="vehicle_id", row_number=index),
                assigned_from=_parse_import_datetime(row.get("assigned_from"), field="assigned_from", row_number=index),
                remarks=row.get("remarks") or None,
            )
        )
        created += 1
    return {"created": created}
